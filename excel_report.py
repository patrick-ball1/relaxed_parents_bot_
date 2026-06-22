"""
excel_report.py — автоматический экспорт данных пользователей в Excel.

Функции:
  • rebuild_excel()       — пересобрать весь файл из БД (при старте/по расписанию)
  • append_user_row()     — добавить/обновить строку пользователя
  • append_event_row()    — добавить строку события в лист «События»

Файл analytics_aptechka.xlsx создаётся / обновляется автоматически.
"""

import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from config import DATABASE_URL
from database import engine, SessionLocal

EXCEL_PATH = os.getenv("EXCEL_PATH", "analytics_aptechka.xlsx")

# ─── Цвета ────────────────────────────────────────────────────────────────────
CLR_HEADER_USERS  = "2E5FA3"   # синий заголовок — лист Пользователи
CLR_HEADER_EVENTS = "1E7E34"   # зелёный заголовок — лист События
CLR_HEADER_FUNNEL = "8B4E00"   # коричневый — лист Воронка
CLR_WHITE         = "FFFFFF"
CLR_ROW_ALT       = "EFF3FB"   # светло-голубой через строку
CLR_HOT_LEAD      = "FFD700"   # золото — горячий лид (day3_interest_clicked)
CLR_PAID          = "C6EFCE"   # зелёный фон — купил

SEGMENT_RU = {
    "pregnant": "Беременна",
    "0_3":      "0–3 мес",
    "3_6":      "3–6 мес",
    "6_12":     "6–12 мес",
    "1plus":    "Больше года",
}

# ─── Стили ────────────────────────────────────────────────────────────────────
def _header_font(color=CLR_WHITE):
    return Font(name="Arial", bold=True, color=color, size=10)

def _cell_font():
    return Font(name="Arial", size=10)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_header_row(ws, headers: list, fill_color: str):
    """Записывает строку заголовков с форматированием."""
    ws.row_dimensions[1].height = 30
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _header_font()
        cell.fill = _fill(fill_color)
        cell.alignment = _center()
        cell.border = _thin_border()


def _style_data_row(ws, row_num: int, num_cols: int, alt: bool = False, bg: Optional[str] = None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _cell_font()
        cell.border = _thin_border()
        cell.alignment = _left()
        if bg:
            cell.fill = _fill(bg)
        elif alt:
            cell.fill = _fill(CLR_ROW_ALT)


# ─── Лист «Пользователи» ──────────────────────────────────────────────────────
USERS_HEADERS = [
    "ID пользователя",
    "Имя",
    "Сегмент",
    "Стадия воронки",
    "Дата старта",
    "Зарегистрирован",
    "Горячий лид",
    "Купил",
]
USERS_COL_WIDTHS = [18, 18, 16, 20, 22, 22, 14, 10]


def _build_users_sheet(ws, rows: list):
    _set_header_row(ws, USERS_HEADERS, CLR_HEADER_USERS)
    for i, w in enumerate(USERS_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r_idx, row in enumerate(rows, 2):
        user_id, name, segment, stage, start_time, created_at, hot_lead, paid = row

        ws.cell(r_idx, 1, value=user_id)
        ws.cell(r_idx, 2, value=name or "—")
        ws.cell(r_idx, 3, value=SEGMENT_RU.get(segment, segment or "—"))
        ws.cell(r_idx, 4, value=stage or "new")
        ws.cell(r_idx, 5, value=_fmt_dt(start_time))
        ws.cell(r_idx, 6, value=_fmt_dt(created_at))
        ws.cell(r_idx, 7, value="✅" if hot_lead else "")
        ws.cell(r_idx, 8, value="✅" if paid else "")

        bg = CLR_PAID if paid else (CLR_HOT_LEAD if hot_lead else None)
        _style_data_row(ws, r_idx, len(USERS_HEADERS), alt=(r_idx % 2 == 0), bg=bg)


# ─── Лист «События» ───────────────────────────────────────────────────────────
EVENTS_HEADERS = [
    "Дата / время",
    "ID пользователя",
    "Имя",
    "Событие",
    "Данные",
]
EVENTS_COL_WIDTHS = [22, 18, 18, 28, 40]


def _build_events_sheet(ws, rows: list):
    _set_header_row(ws, EVENTS_HEADERS, CLR_HEADER_EVENTS)
    for i, w in enumerate(EVENTS_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r_idx, row in enumerate(rows, 2):
        created_at, user_id, name, event, data = row
        ws.cell(r_idx, 1, value=_fmt_dt(created_at))
        ws.cell(r_idx, 2, value=user_id)
        ws.cell(r_idx, 3, value=name or "—")
        ws.cell(r_idx, 4, value=event)
        ws.cell(r_idx, 5, value=data or "")
        _style_data_row(ws, r_idx, len(EVENTS_HEADERS), alt=(r_idx % 2 == 0))


# ─── Лист «Воронка» ───────────────────────────────────────────────────────────
FUNNEL_HEADERS = ["Этап воронки", "Кол-во пользователей", "% от старта"]
FUNNEL_STEPS = [
    ("bot_started",           "🚀 Зашёл в бот"),
    ("webinar_offered",       "🎬 Увидел кнопку вебинара"),
    ("webinar_watched",       "▶️ Нажал смотреть вебинар"),
    ("thank_you_sent",        "🙏 Получил благодарность"),
    ("pitch_academy_sent",    "✨ Получил питч академии"),
    ("academy_link_clicked",  "🎓 Нажал на кнопку академии"),
    ("offer_3h_sent",         "💰 Получил оффер 4980₸"),
    ("reviews_6h_sent",       "💬 Получил отзывы"),
    ("story_julia_sent",      "📖 Получил историю Юлии"),
    ("final_message_sent",    "🏁 Получил финальный призыв"),
]


def _build_funnel_sheet(ws, counts: dict):
    _set_header_row(ws, FUNNEL_HEADERS, CLR_HEADER_FUNNEL)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16

    total = counts.get("bot_started", 0) or 1  # избегаем деления на 0

    for r_idx, (event, label) in enumerate(FUNNEL_STEPS, 2):
        cnt = counts.get(event, 0)
        pct = round(cnt / total * 100, 1)
        ws.cell(r_idx, 1, value=label)
        ws.cell(r_idx, 2, value=cnt)
        ws.cell(r_idx, 3, value=f"{pct}%")
        _style_data_row(ws, r_idx, 3, alt=(r_idx % 2 == 0))

    # Итоговая строка
    last = len(FUNNEL_STEPS) + 2
    ws.cell(last, 1, value="Всего в базе").font = Font(name="Arial", bold=True, size=10)
    ws.cell(last, 2, value=total).font = Font(name="Arial", bold=True, size=10)


# ─── Вспомогательные функции ──────────────────────────────────────────────────
def _fmt_dt(value) -> str:
    if not value:
        return "—"
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except ValueError:
            return value
    return value.strftime("%d.%m.%Y %H:%M")


# ─── Основные публичные функции ───────────────────────────────────────────────
async def rebuild_excel():
    """Полностью пересобирает Excel из базы данных. Вызывается при старте и по расписанию."""
    async with SessionLocal() as s:
        # Пользователи
        users_q = await s.execute(text("""
            SELECT
                u.user_id,
                u.name,
                u.segment,
                u.stage,
                u.start_time,
                u.created_at,
                CASE WHEN EXISTS (
                    SELECT 1 FROM analytics a
                    WHERE a.user_id = u.user_id AND a.event = 'day3_interest_clicked'
                ) THEN 1 ELSE 0 END AS hot_lead,
                CASE WHEN u.stage = 'paid' THEN 1 ELSE 0 END AS paid
            FROM users u
            ORDER BY u.created_at DESC
        """))
        users_rows = users_q.fetchall()

        # События (последние 5000)
        events_q = await s.execute(text("""
            SELECT a.created_at, a.user_id, u.name, a.event, a.data
            FROM analytics a
            LEFT JOIN users u ON u.user_id = a.user_id
            ORDER BY a.created_at DESC
            LIMIT 5000
        """))
        events_rows = events_q.fetchall()

        # Воронка — счётчики
        funnel_events = [e for e, _ in FUNNEL_STEPS]
        counts = {}
        for ev in funnel_events:
            r = await s.execute(
                text("SELECT COUNT(DISTINCT user_id) FROM analytics WHERE event=:ev"),
                {"ev": ev},
            )
            counts[ev] = r.scalar() or 0

    wb = Workbook()

    # Лист 1: Пользователи
    ws_users = wb.active
    ws_users.title = "👥 Пользователи"
    ws_users.freeze_panes = "A2"
    _build_users_sheet(ws_users, users_rows)

    # Лист 2: Воронка
    ws_funnel = wb.create_sheet("📊 Воронка")
    ws_funnel.freeze_panes = "A2"
    _build_funnel_sheet(ws_funnel, counts)

    # Лист 3: События
    ws_events = wb.create_sheet("📋 События")
    ws_events.freeze_panes = "A2"
    _build_events_sheet(ws_events, events_rows)

    # Метаданные
    wb.properties.title = "Аналитика бота Аптечка"
    wb.properties.creator = "Aptechka Bot"

    wb.save(EXCEL_PATH)


async def refresh_excel_user(user_id: int):
    """
    Быстрое обновление одной строки пользователя в листе «Пользователи».
    Если строка не найдена — полностью перестраивает файл.
    """
    async with SessionLocal() as s:
        r = await s.execute(text("""
            SELECT
                u.user_id, u.name, u.segment, u.stage, u.start_time, u.created_at,
                CASE WHEN EXISTS (
                    SELECT 1 FROM analytics a
                    WHERE a.user_id = u.user_id AND a.event = 'day3_interest_clicked'
                ) THEN 1 ELSE 0 END,
                CASE WHEN u.stage = 'paid' THEN 1 ELSE 0 END
            FROM users u WHERE u.user_id = :uid
        """), {"uid": user_id})
        row = r.fetchone()

    if not row or not os.path.exists(EXCEL_PATH):
        await rebuild_excel()
        return

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["👥 Пользователи"]

        # Ищем строку пользователя
        target_row = None
        for r_idx in range(2, ws.max_row + 1):
            if ws.cell(r_idx, 1).value == user_id:
                target_row = r_idx
                break

        user_id_, name, segment, stage, start_time, created_at, hot_lead, paid = row

        if target_row is None:
            target_row = ws.max_row + 1

        ws.cell(target_row, 1, value=user_id_)
        ws.cell(target_row, 2, value=name or "—")
        ws.cell(target_row, 3, value=SEGMENT_RU.get(segment, segment or "—"))
        ws.cell(target_row, 4, value=stage or "new")
        ws.cell(target_row, 5, value=_fmt_dt(start_time))
        ws.cell(target_row, 6, value=_fmt_dt(created_at))
        ws.cell(target_row, 7, value="✅" if hot_lead else "")
        ws.cell(target_row, 8, value="✅" if paid else "")

        bg = CLR_PAID if paid else (CLR_HOT_LEAD if hot_lead else None)
        _style_data_row(ws, target_row, len(USERS_HEADERS), alt=(target_row % 2 == 0), bg=bg)

        wb.save(EXCEL_PATH)
    except Exception:
        # Если файл повреждён — перестраиваем полностью
        await rebuild_excel()


async def append_event_to_excel(user_id: int, name: str, event: str, data: str = ""):
    """Добавляет строку события в лист «События» без полной перестройки файла."""
    if not os.path.exists(EXCEL_PATH):
        await rebuild_excel()
        return

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["📋 События"]
        new_row = ws.max_row + 1
        now = datetime.utcnow().strftime("%d.%m.%Y %H:%M")
        ws.cell(new_row, 1, value=now)
        ws.cell(new_row, 2, value=user_id)
        ws.cell(new_row, 3, value=name or "—")
        ws.cell(new_row, 4, value=event)
        ws.cell(new_row, 5, value=data)
        _style_data_row(ws, new_row, len(EVENTS_HEADERS), alt=(new_row % 2 == 0))
        wb.save(EXCEL_PATH)
    except Exception:
        await rebuild_excel()
